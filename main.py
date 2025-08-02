'''
using for delays in the web scraping.
'''
import time
import sys
import requests

from openpyxl import load_workbook
from bs4 import BeautifulSoup
from bs4 import Tag

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

service = Service(ChromeDriverManager().install())


def get_element_inner_html(url, class_name):
    '''
    Gets an element given the class and the url.
    '''
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')

        elements = soup.find_all(class_=class_name)

        if elements:
            el = elements[0]
            if isinstance(el, Tag):
                return el.decode_contents()

            return str(el)

        return f"No element found with ID '{class_name};"

    except requests.RequestException as e:
        return f"Error fetching the webpage: {e}"


def get_price(url):
    '''
    Returns the price of a house given it's url.
    '''
    return str(num_string_to_num(get_element_inner_html(url, 'display_price')))


def get_homes_url(address):
    '''
    Gets the url of a house, given it's address.
    '''
    # Set up headless Chrome
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')

    driver = webdriver.Chrome(service=service, options=options)

    try:
        # Step 1: Go to the homes.co.nz homepage
        driver.get("https://homes.co.nz")
        wait = WebDriverWait(driver, 10)
        search_input = wait.until(EC.presence_of_element_located((By.ID, "autocomplete-search")))

        # Alternative approach: Use JavaScript to set the value directly
        # This bypasses any JavaScript that might be clearing the input
        address_str = str(address).strip()

        # Method 1: Try normal typing first
        search_input.clear()
        search_input.send_keys(address_str)
        time.sleep(1)

        # Find and click the search button
        try:
            selector = '.searchButton .homes-button-main'
            search_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector)))
            search_button.click()
        except Exception as e:
            print(f"Error clicking search button: {e}")
            # Fallback to Enter key
            # print("Falling back to Enter key")
            search_input.send_keys(Keys.RETURN)

        # Wait for results to load
        # time.sleep(5)

        selector = 'a.heroImage.ng-star-inserted'

        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, selector))
            )
            results = driver.find_elements(By.CSS_SELECTOR, selector)
            print(results)

            if not results:
                return None

            url = results[0].get_attribute('href')
            print(url)
            return url
        except Exception as e:
            print(e)
            return None
    finally:
        driver.quit()


def num_string_to_num(string):
    '''
    Takes a number string like 543k, or 1.5M, and returns a number like 543000, or 1500000
    '''
    try:
        num = float(string[0:-1])
        if not num:
            return "NaN"
        letter = string[-1:]
        if letter == "K":
            num *= 1000
        elif letter == "M":
            num *= 1000000
        return num
    except ValueError:
        return ""


def main():
    '''
    The main method, takes 1 system argument, for the path to the excel file.
    '''
    if len(sys.argv) < 2:
        print("no path provided")
        return

    wb = load_workbook(sys.argv[1])

    ws = wb["Sheet1"]

    for i in range(1, ws.max_row):
        cell_value = ws[f"B{i}"].value

        url = ""
        price = ""
        if cell_value is None:
            print(f"Row {i} is empty — skipping.")
        else:
            url = ws[f"C{i}"].value
            if url == "" or url == "None" or url is None:
                url = get_homes_url(cell_value)
                ws[f"C{i}"] = url

            print(url)
            price = get_price(url)
            if "None" in price:
                price = ""
            if "Invalid" in price:
                price = ""
            if "element" in price:
                price = ""
        try:
            ws["D" + str(i)] = float(price)
        except ValueError:
            ws["D" + str(i)] = ""

    wb.save(sys.argv[1])


main()
